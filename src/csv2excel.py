"""
module: This module loads a given csv file and converts it into a json
"""
from pathlib import Path
from openpyxl import load_workbook
import csvfile

from src.mappings.mapper import Map, KeysGroup

def build(fname: str):
    """
    loads given csv file into object by calling __load_csv().
    """
    try:
        # cj = Map({"paid fte": "one", "effective date": "yyyy-mm-dd"}, KeysGroup.JOB_CHANGE)
        # print(cj.wdmap.get("Paid FTE", None))
        records = []
        rows = __load_csv(fname)
        for row in rows:
            records.append(Map(row, KeysGroup.JOB_CHANGE))

        if len(records) == 0:
            print("no records found, aborting creating excel file.")
            return

        __print_log(f'loaded {len(records)} records from {fname}')

        __copy_to_excel(fname, records)
        # __write_to_excel(fname, records)
        __print_log("Successfully created xlsx")

    except Exception as error:
        print(error.args)

def __load_csv(fname: str):
    """
        loads csv file and converts it into a list of objects
        e.g
        [
            {"key1": "{row1.col1}", "key2": "{row1.col1}"},
            {"key1": "{row2.col1}", "key2": "{row2.col1}"},
            ... 
        ]
    """
    rows = csvfile.load(fname)
    return rows

def __copy_to_excel(fname: str, records: list):
    """
        updates Change_job_sample.xlsx's Change Job sheet and writes into {fname}-output.xlsx.
        input: fname: str
        records: list of instance of Map
    """

    if len(records) == 0:
        return False

    workbook = load_workbook(filename = './static/Change_job_sample.xlsx')
    worksheet = workbook['Change Job']
    header = worksheet[5]
    data_row = 6 #data cell starts on the 6th row after header finishes off in Chane_job_sample.xlsx

    for record in records:
        for header_cell in header:
            value = record.wdmap.get(header_cell.value, None)

            if value is None:
                continue

            worksheet.cell(row=data_row,column=header_cell.column).value = value
        data_row += 1

    workbook.save(__xlsx_fname(fname))

def __xlsx_fname(ref_fname: str):
    fname = Path(ref_fname).stem
    return Path(fname+"-output").with_suffix('.xlsx')

def __print_log(*logs):
    text = ""
    for log in logs:
        text += str(log)
    print("*********", text, "*********")
