from pathlib import Path

from openpyxl import load_workbook

from src.constants import (
    BusinessProcessType,
    BusinessProcessTypeToName,
    BusinessProcessTypeToProcessInstruction,
)
from src.eib_builder_logger import EibBuilderLogger
from src.mappings.mapper import Map

logger = EibBuilderLogger()


class EibBuilder:
    DATA_ROW_START = 6  # data cell starts on the 6th row after header finishes off in Chane_job_sample.xlsx

    def __init__(self):
        pass

    @classmethod
    def build(
        cls,
        data_records: list[Map],
        output_file_name: str = "change_job_eib",
        business_process_types: list[BusinessProcessType] = [
            BusinessProcessType.CHANGE_JOB
        ],
    ) -> None:
        """
        loads given csv file into object by calling __load_csv().
        """
        try:
            if not data_records or len(data_records) < 1:
                error_message = "No input records provided to convert to EIB. Please provide valid data records."
                logger.log_error(error_message)
                return
            build_eib = EibBuilder()
            workbook = load_workbook(
                filename="./templates/template_change_job_v42.1.xlsx"
            )
            for process_type in business_process_types:
                build_eib.populate_data(workbook, data_records, process_type)

            workbook.save(build_eib.generate_output_filename(output_file_name))

            logger.log_messages(["Successfully created xlsx"])

        except Exception as error:
            logger.log_error(f"Error occurred: {error}")

    def populate_data(
        self, workbook, records: list[Map], process_type: BusinessProcessType
    ) -> None:
        worksheet = workbook[BusinessProcessTypeToName[process_type]]
        excel_header = worksheet and worksheet[5]
        data_row = EibBuilder.DATA_ROW_START
        row_count = 1

        for record in records:
            mappings = record.wdmap(process_type)
            self.populate_excel(worksheet, mappings, row_count, excel_header, data_row)
            data_row += 1
        # set business process parameter for change job
        worksheet_overview = workbook["Overview"]
        worksheet_overview.cell(
            row=BusinessProcessTypeToProcessInstruction[process_type].get("row"),
            column=3,
        ).value = BusinessProcessTypeToProcessInstruction[process_type].get("value")

    def populate_excel(self, work_sheet, mappings, row_count, excel_header, data_row):
        for header_cell in excel_header:
            if header_cell.column_letter == "B":
                work_sheet.cell(
                    row=data_row, column=header_cell.column
                ).value = row_count
                continue
            value = mappings.get(header_cell.column_letter, None)

            if value is None:
                continue

            work_sheet.cell(row=data_row, column=header_cell.column).value = value

    def generate_output_filename(self, ref_fname: str) -> str:
        """
        Generates output filename with appropriate suffix.
        """
        fname = Path(ref_fname).stem
        return str(Path(fname + "-output").with_suffix(".xlsx"))
